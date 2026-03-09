from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, jsonify
from datetime import date, datetime, timedelta
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from dotenv import load_dotenv
import os
from sqlalchemy import func, desc

# ── Database imports ─────────────────────────────────────────────
from database import SessionLocal, init_db, Labour, AttendanceRecord, CategoryScore, FeedbackViolation, StrikeEvent

load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET_KEY", "super-secret-key-123xyz-change-this-in-production")

# Initialize database tables
with app.app_context():
    init_db()

# Hardcoded credentials (move to DB + hashed later)
ADMIN_USER = "User"
ADMIN_PASS = "1234"

# Feedback categories and items (move to DB later)
FEEDBACK_STRUCTURE = {
    "Respectful communications": [
        "Always communicate respectfully and professionally with both customers and staff",
        "Handle conflicts in a calm, fair, and constructive manner",
        "Be Solution-Oriented, Not Blame-Oriented",
        "Practice active listening",
        "Maintain honesty and transparency in all actions and communications"
    ],
    "Etiquette": [
        "Use mobile phones only for work-related purposes during shift hours",
        "Take prior permission before going for breaks or stepping out in working hours",
        "Always keep your workplace clean and organized",
        "Follow company rules, reporting formats, and SOPs consistently without the need for constant reminders"
    ],
    "Decorum/digital decorum": [
        "Maintain punctuality by arriving and leaving as per designated working hours",
        "Maintain dress code (T shirt)",
        "Break times are fixed as per HR policy and extra breaks will invite penalties"
    ],
    "Harm to company brand and reputation, revenues": [
        "Theft from company is a serious offence and can result in termination",
        "Badmouthing the company is strictly forbidden amongst employees, especially in public places"
    ],
    "Teamwork": [
        "No quarrels or arguments with co-workers",
        "No physical fights with co-workers",
        "Always cooperate with manager and co-workers"
    ]
}
CATEGORY_NAMES = list(FEEDBACK_STRUCTURE.keys())


def login_required(f):
    def wrapper(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    wrapper.__name__ = f.__name__
    return wrapper


@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")
        if username == ADMIN_USER and password == ADMIN_PASS:
            session["logged_in"] = True
            flash("Login successful", "success")
            return redirect(url_for("dashboard"))
        else:
            flash("Wrong username or password", "danger")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.pop("logged_in", None)
    flash("Logged out successfully", "info")
    return redirect(url_for("login"))


@app.route("/dashboard")
@login_required
def dashboard():
    db = SessionLocal()
    try:
        # Load active labours
        labours_db = db.query(Labour).filter_by(is_active=True).order_by(Labour.id).all()
        labours = [
            {"id": l.id, "name": l.name, "surname": l.surname, "phone": l.phone}
            for l in labours_db
        ]

        today_iso = date.today().isoformat()

        # Load today's attendance records
        today_records = db.query(AttendanceRecord).filter(
            AttendanceRecord.record_date == today_iso
        ).all()

        today_data = {}
        for rec in today_records:
            today_data[rec.labour_id] = {
                "status": rec.status,
                "advance": float(rec.advance_amount) if rec.advance_amount else None,
                "total_points": rec.total_points,
                "total_break_minutes": rec.total_break_minutes or 0,
                "category_scores": {},
                "feedback": {},
                "breaks": [],
                "frozen": rec.frozen  # to disable status if frozen
            }

        # Load category scores for today
        scores = db.query(CategoryScore).join(AttendanceRecord).filter(
            AttendanceRecord.record_date == today_iso
        ).all()
        for score in scores:
            labour_id = db.query(AttendanceRecord.labour_id).filter_by(id=score.attendance_record_id).scalar()
            if labour_id and labour_id in today_data:
                today_data[labour_id]["category_scores"][score.category_id] = score.score

        # Load feedback violations for today
        violations_today = db.query(FeedbackViolation).filter(
            FeedbackViolation.violation_date == today_iso
        ).all()
        for v in violations_today:
            lid = v.labour_id
            if lid in today_data:
                today_data[lid]["feedback"][v.checklist_item_id] = True

        # Load history (filtered)
        filter_labour_id = request.args.get('filter_labour', 'all', type=str)
        filter_from = request.args.get('from_date', '')
        filter_to = request.args.get('to_date', '')

        if filter_from == '' and filter_to == '':
            to_date = date.today()
            from_date = to_date - timedelta(days=30)
            filter_from = from_date.isoformat()
            filter_to = to_date.isoformat()

        try:
            from_date_obj = datetime.strptime(filter_from, '%Y-%m-%d').date() if filter_from else None
            to_date_obj = datetime.strptime(filter_to, '%Y-%m-%d').date() if filter_to else None
        except:
            from_date_obj = None
            to_date_obj = None
            flash("Invalid date format.", "danger")

        history_query = db.query(AttendanceRecord).filter(
            AttendanceRecord.record_date >= (from_date_obj or date.today() - timedelta(days=30)),
            AttendanceRecord.record_date <= (to_date_obj or date.today())
        ).order_by(desc(AttendanceRecord.record_date))

        if filter_labour_id != 'all':
            history_query = history_query.filter(AttendanceRecord.labour_id == int(filter_labour_id))

        history_records = history_query.all()

        filtered_history = {}
        for rec in history_records:
            date_str = rec.record_date.isoformat()
            if date_str not in filtered_history:
                filtered_history[date_str] = {}
            filtered_history[date_str][rec.labour_id] = {
                "status": rec.status,
                "advance": float(rec.advance_amount) if rec.advance_amount else 0,
                "total_points": rec.total_points,
                "total_break_minutes": rec.total_break_minutes or 0,
                "category_scores": {},  # can load if needed
                "feedback": {},         # can load if needed
                "breaks": [],
                "frozen": rec.frozen
            }

        # Load strikes
        labour_strikes = {}
        for labour in labours:
            lid = labour["id"]
            count = db.query(func.count(FeedbackViolation.id)).filter(
                FeedbackViolation.labour_id == lid
            ).scalar() or 0
            total_strikes = count // 3
            labour_strikes[lid] = {'total': total_strikes, 'details': []}

    finally:
        db.close()

    return render_template(
        "dashboard.html",
        labours=labours,
        today_data=today_data,
        today_date=date.today().strftime("%d %B %Y"),
        today_iso=today_iso,
        history=filtered_history,
        feedback_structure=FEEDBACK_STRUCTURE,
        filter_labour=filter_labour_id,
        filter_from=filter_from,
        filter_to=filter_to,
        category_names=CATEGORY_NAMES,
        labour_strikes=labour_strikes,
        datetime=datetime
    )


@app.route("/add_labour", methods=["POST"])
@login_required
def add_labour():
    name = request.form.get("name", "").strip()
    surname = request.form.get("surname", "").strip()
    phone = request.form.get("phone", "").strip()

    if not name or not surname:
        flash("Name and Surname are required", "danger")
        return redirect(url_for("dashboard"))

    db = SessionLocal()
    try:
        existing = db.query(Labour).filter_by(name=name, surname=surname).first()
        if existing:
            flash(f"Labour {name} {surname} already exists!", "warning")
            return redirect(url_for("dashboard"))

        new_labour = Labour(name=name, surname=surname, phone=phone or None)
        db.add(new_labour)
        db.commit()
        flash(f"Labour {name} {surname} added successfully (ID: {new_labour.id})", "success")
    except Exception as e:
        db.rollback()
        flash(f"Database error: {str(e)}", "danger")
    finally:
        db.close()

    return redirect(url_for("dashboard"))


@app.route("/mark_attendance", methods=["POST"])
@login_required
def mark_attendance():
    today_iso = date.today().isoformat()
    db = SessionLocal()
    try:
        updated_lids = set()
        for key in request.form:
            if key.startswith(("status_", "advance_", "score_", "fb_")):
                try:
                    lid = int(key.split("_")[1])
                    updated_lids.add(lid)
                except:
                    pass

        for lid in updated_lids:
            record = db.query(AttendanceRecord).filter_by(
                labour_id=lid, record_date=today_iso
            ).first()

            if not record:
                record = AttendanceRecord(
                    labour_id=lid,
                    record_date=today_iso,
                    status="Present",
                    advance_amount=0,
                    total_points=0,
                    total_break_minutes=0,
                    frozen=False  # will set frozen=True after save if needed
                )
                db.add(record)
                db.flush()

            # Update status
            status_key = f"status_{lid}"
            if status_key in request.form:
                status = request.form[status_key]
                if status in ["Present", "Absent", "Half Day"]:
                    record.status = status

            # Update advance
            advance_key = f"advance_{lid}"
            if advance_key in request.form:
                try:
                    advance = float(request.form[advance_key].strip() or 0)
                    record.advance_amount = advance
                except:
                    pass

            # Update scores
            record.total_points = 0
            for key, value in request.form.items():
                if key.startswith(f"score_{lid}_"):
                    try:
                        cat_idx = int(key.split("_")[2])
                        score_val = int(value)
                        score_rec = db.query(CategoryScore).filter_by(
                            attendance_record_id=record.id,
                            category_id=cat_idx
                        ).first()
                        if not score_rec:
                            score_rec = CategoryScore(
                                attendance_record_id=record.id,
                                category_id=cat_idx,
                                score=score_val
                            )
                            db.add(score_rec)
                        else:
                            score_rec.score = score_val
                        record.total_points += score_val
                    except:
                        pass

            # Update feedback violations
            checked_items = set()
            for key in request.form:
                if key.startswith(f"fb_{lid}_"):
                    try:
                        parts = key.split("_")
                        item_idx = int(parts[2])
                        cat_idx = int(parts[3])
                        item_id = cat_idx * 10 + item_idx  # dummy mapping
                        checked_items.add(item_id)
                    except:
                        pass

            db.query(FeedbackViolation).filter_by(attendance_record_id=record.id).delete()

            for item_id in checked_items:
                violation = FeedbackViolation(
                    attendance_record_id=record.id,
                    labour_id=lid,
                    checklist_item_id=item_id,
                    violation_date=today_iso
                )
                db.add(violation)

                count = db.query(func.count(FeedbackViolation.id)).filter(
                    FeedbackViolation.labour_id == lid,
                    FeedbackViolation.checklist_item_id == item_id
                ).scalar() or 0

                if count % 3 == 0 and count > 0:
                    strike = StrikeEvent(
                        labour_id=lid,
                        checklist_item_id=item_id,
                        record_date=today_iso,
                        strike_number=count // 3,
                        attendance_record_id=record.id
                    )
                    db.add(strike)

        # Freeze status only after save (set frozen=True)
        for lid in updated_lids:
            record = db.query(AttendanceRecord).filter_by(labour_id=lid, record_date=today_iso).first()
            if record:
                record.frozen = True  # Freeze status for this day

        db.commit()
        flash("Today's attendance saved successfully (Status frozen for today)", "success")
    except Exception as e:
        db.rollback()
        flash(f"Error saving attendance: {str(e)}", "danger")
    finally:
        db.close()

    return redirect(url_for("dashboard"))


@app.route("/toggle_break", methods=["POST"])
@login_required
def toggle_break():
    today_iso = date.today().isoformat()
    data = request.get_json()
    labour_id = int(data["labour_id"])
    is_on = data["is_on"]

    print(f"[{datetime.now().strftime('%H:%M:%S')}] Break toggle request for labour {labour_id} - Action: {'ON' if is_on else 'OFF'}")

    db = SessionLocal()
    try:
        record = db.query(AttendanceRecord).filter_by(
            labour_id=labour_id, record_date=today_iso
        ).first()

        if not record:
            print(f"[{datetime.now().strftime('%H:%M:%S')}] ERROR: No attendance record found for labour {labour_id} on {today_iso}")
            return jsonify({"success": False, "message": "No attendance record found for today"})

        print(f"[{datetime.now().strftime('%H:%M:%S')}] Record found - Current total_break_minutes: {record.total_break_minutes or 0}")
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Current break start in DB: {record.current_break_start}")

        now = datetime.now()

        if is_on:
            # Start break - save to REAL column
            record.current_break_start = now
            print(f"[{datetime.now().strftime('%H:%M:%S')}] Break STARTED - Start time saved to DB column: {now}")
            message = "Break started"
            total_break = record.total_break_minutes or 0
        else:
            # End break - use REAL column
            if record.current_break_start:
                duration = int((now - record.current_break_start).total_seconds() / 60)
                print(f"[{datetime.now().strftime('%H:%M:%S')}] Break ENDED - Duration calculated: {duration} min (from {record.current_break_start} to {now})")
                record.total_break_minutes = (record.total_break_minutes or 0) + duration
                record.current_break_start = None  # Clear the start time
                print(f"[{datetime.now().strftime('%H:%M:%S')}] New total saved: {record.total_break_minutes}")
                message = f"Break ended ({duration} min)"
            else:
                print(f"[{datetime.now().strftime('%H:%M:%S')}] No active break found to end (current_break_start is NULL)")
                message = "No active break to end"
            total_break = record.total_break_minutes or 0

        db.commit()
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Commit successful - Returning total_break: {total_break}")
        return jsonify({"success": True, "message": message, "total_break": total_break})
    except Exception as e:
        db.rollback()
        print(f"[{datetime.now().strftime('%H:%M:%S')}] ERROR during processing: {str(e)}")
        return jsonify({"success": False, "message": f"Error: {str(e)}"})
    finally:
        db.close()

@app.route("/export_report")
@login_required
def export_report():
    filter_labour_id = request.args.get('filter_labour', 'all')
    filter_from = request.args.get('from_date', '')
    filter_to = request.args.get('to_date', '')
    export_format = request.args.get('format', 'excel')

    try:
        from_date = datetime.strptime(filter_from, '%Y-%m-%d').date() if filter_from else None
        to_date = datetime.strptime(filter_to, '%Y-%m-%d').date() if filter_to else None
    except:
        flash("Invalid date format", "danger")
        return redirect(url_for("dashboard"))

    db = SessionLocal()
    try:
        labours_db = {l.id: l for l in db.query(Labour).all()}

        query = db.query(AttendanceRecord).order_by(AttendanceRecord.record_date.desc())
        if from_date:
            query = query.filter(AttendanceRecord.record_date >= from_date)
        if to_date:
            query = query.filter(AttendanceRecord.record_date <= to_date)
        if filter_labour_id != 'all':
            query = query.filter(AttendanceRecord.labour_id == int(filter_labour_id))

        records = query.all()

        filtered_data = []
        for rec in records:
            labour = labours_db.get(rec.labour_id)
            if not labour:
                continue

            row = {
                'date': rec.record_date.isoformat(),
                'labour_id': rec.labour_id,
                'name': labour.name,
                'surname': labour.surname,
                'phone': labour.phone or '',
                'status': rec.status or '',
                'advance': float(rec.advance_amount or 0),
                'total_points': rec.total_points or 0,
                'total_break_minutes': rec.total_break_minutes or 0,
                'strike_details': "Strikes from DB",
                'strike_count': 0,
                'checklist_details': "Feedback from DB"
            }
            filtered_data.append(row)

    finally:
        db.close()

    filtered_data.sort(key=lambda x: x['date'], reverse=True)

    if export_format == 'csv':
        return export_to_csv(filtered_data)
    else:
        return export_to_excel(filtered_data)


def export_to_csv(data):
    if not data:
        flash("No data to export", "warning")
        return redirect(url_for("dashboard"))

    output = io.StringIO()
    fieldnames = ['date', 'name', 'surname', 'phone', 'status', 'advance', 
                  'total_points', 'total_break_minutes', 'strike_details', 'checklist_details']
    for idx in range(1, len(CATEGORY_NAMES) + 1):
        fieldnames.append(f'score_cat{idx}')

    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()

    for row in data:
        writer.writerow(row)

    output.seek(0)
    filename = f"attendance_report_{date.today().isoformat()}.csv"

    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        mimetype='text/csv',
        as_attachment=True,
        download_name=filename
    )


def export_to_excel(data):
    if not data:
        flash("No data to export", "warning")
        return redirect(url_for("dashboard"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_alignment = Alignment(horizontal="center")

    headers = ['Date', 'Name', 'Surname', 'Phone', 'Status', 'Advance (₹)', 
               'Total Points', 'Total Break (min)', 'Strike Details', 'Checklist Details']
    for idx, cat_name in enumerate(CATEGORY_NAMES, 1):
        headers.append(f'{cat_name} Score')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    for row_idx, row_data in enumerate(data, 2):
        ws.cell(row=row_idx, column=1, value=row_data['date'])
        ws.cell(row=row_idx, column=2, value=row_data['name'])
        ws.cell(row=row_idx, column=3, value=row_data['surname'])
        ws.cell(row=row_idx, column=4, value=row_data['phone'])
        ws.cell(row=row_idx, column=5, value=row_data['status'])
        ws.cell(row=row_idx, column=6, value=row_data['advance'])
        ws.cell(row=row_idx, column=7, value=row_data['total_points'])
        ws.cell(row=row_idx, column=8, value=row_data['total_break_minutes'])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"attendance_report_{date.today().isoformat()}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )



# Production-ready: Railway uses PORT env var, turns off debug
if __name__ == "__main__":
    # Get PORT from Railway (fallback to 5000 for local)
    port = int(os.environ.get("PORT", 5000))
    # Disable debug in production (Railway sets no FLASK_DEBUG)
    debug = os.environ.get("FLASK_DEBUG", "False").lower() == "true"
    app.run(host="0.0.0.0", port=port, debug=debug)
